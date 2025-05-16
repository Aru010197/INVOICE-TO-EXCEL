import pytesseract
from PIL import Image
import os
import openpyxl
import re
import PyPDF2  # Added for PDF text extraction
from pdf2image import convert_from_path  # Added for PDF to image conversion

# --- Configuration ---
# If Tesseract is not in your PATH, you'll need to set the tesseract_cmd
# Example for macOS if installed via Homebrew:
pytesseract.pytesseract.tesseract_cmd = r'/opt/homebrew/bin/tesseract'
# Ensure poppler is installed for pdf2image: brew install poppler
# Example for Windows:
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
# For pdf2image on Windows, you might need to provide the poppler path to convert_from_path
# poppler_path = r"C:\path\to\poppler-xx.xx.x\bin"

INVOICE_DIR = "invoices"  # Relative path to the invoices folder
OUTPUT_EXCEL_FILE = "invoice_data.xlsx"  # Relative path for the output Excel file

# --- Helper Functions ---

def ocr_image(image_path):
    """
    Performs OCR on a given image file.
    Returns the extracted text.
    """
    try:
        text = pytesseract.image_to_string(Image.open(image_path))
        print(f"Successfully OCR'd image: {image_path}")
        return text
    except Exception as e:
        print(f"Error during OCR for image {image_path}: {e}")
        return ""

def extract_text_from_pdf(pdf_path):
    """
    Extracts text directly from a PDF file.
    """
    try:
        with open(pdf_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            text = ""
            for page_num in range(len(reader.pages)):
                page = reader.pages[page_num]
                extracted_page_text = page.extract_text()
                if extracted_page_text:
                    text += extracted_page_text + "\n"
        print(f"Successfully extracted text directly from PDF: {pdf_path}")
        return text
    except Exception as e:
        print(f"Error extracting text directly from PDF {pdf_path}: {e}")
        return ""

def ocr_pdf_as_images(pdf_path, poppler_path=None):
    """
    Converts PDF pages to images and then performs OCR on them.
    Returns concatenated text from all pages.
    """
    try:
        images = convert_from_path(pdf_path, poppler_path=poppler_path)
        combined_text = ""
        for i, image in enumerate(images):
            print(f"OCR'ing page {i+1} of PDF {pdf_path}")
            page_text = pytesseract.image_to_string(image)
            combined_text += page_text + "\n"
        print(f"Successfully OCR'd PDF as images: {pdf_path}")
        return combined_text
    except Exception as e:
        print(f"Error during OCR for PDF {pdf_path} (as images): {e}")
        print("Make sure Poppler is installed and in your PATH, or poppler_path is set correctly in the script.")
        print("On macOS, try: brew install poppler")
        print("On Linux, try: sudo apt-get install poppler-utils")
        return ""

def read_text_from_txt(txt_path):
    """
    Reads text directly from a .txt file.
    """
    try:
        with open(txt_path, 'r', encoding='utf-8') as f:
            text = f.read()
        print(f"Successfully read text from: {txt_path}")
        return text
    except Exception as e:
        print(f"Error reading text file {txt_path}: {e}")
        return ""

def parse_invoice_text(text):
    """
    Parses the OCR'd text to extract invoice details.
    This function attempts to handle variations in field names.
    """
    invoice_data = {
        "invoice_number": None,
        "invoice_date": None,
        "total_amount": None,
        "vendor_name": None, # Or Client Name
        # Add more fields as needed
    }

    lines = text.split('\n')
    text_lower = text.lower() # For case-insensitive keyword searching

    # --- Invoice Number --- (Usually more consistent)
    # Keywords: Invoice No, Invoice #, Invoice Number, Tax Invoice
    # Pattern: Keyword followed by a value (alphanumeric, dashes, etc.)
    invoice_no_patterns = [
        r"invoice\s*(?:no|number|#|id)[:.\s]*([A-Za-z0-9-]+)",
        r"tax\s*invoice[:.\s]*([A-Za-z0-9-]+)"
    ]
    for pattern in invoice_no_patterns:
        match = re.search(pattern, text_lower)
        if match:
            # Extract from original text to preserve case if needed, though usually numbers
            original_match = re.search(pattern, text, re.IGNORECASE)
            if original_match:
                invoice_data["invoice_number"] = original_match.group(1).strip()
                break

    # --- Invoice Date --- (Usually more consistent)
    # Keywords: Date, Invoice Date, Billing Date
    # Pattern: DD/MM/YYYY, MM/DD/YYYY, DD Mon YYYY, etc.
    date_patterns = [
        r"date[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
        r"invoice\s*date[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
        r"billing\s*date[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
        r"(\d{1,2}\s(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[,.]?\s*\d{2,4})",
        r"(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s*\d{1,2}[,.]?\s*\d{2,4}"
    ]
    for pattern in date_patterns:
        match = re.search(pattern, text, re.IGNORECASE) # Dates can have mixed case (e.g. Jan)
        if match:
            invoice_data["invoice_date"] = match.group(1).strip()
            break

    # --- Total Amount --- (Highly variable keywords)
    # Keywords: Total, Amount Due, Balance Due, Grand Total, Net Total, Payable Amount, Total Invoice Amount
    # Pattern: Keyword followed by a currency symbol (optional) and a number (with commas, decimals)
    amount_keywords = [
        "total amount", "amount due", "balance due", "grand total", "net total", 
        "total payable", "invoice total", "total invoice amount", "total", "amount"
    ]
    # Regex to capture amounts like: $1,234.56, 1234.56, 1.234,56 (some European formats)
    amount_regex = r"\$?([\d,]+(?:\.\d{2})?|\d+(?:\.\d{3})*(?:,\d{2})?)" 
    # More robust amount regex: handles optional currency, commas/periods as separators
    # Looks for a keyword and then tries to find an amount on the same line or next few lines.
    
    possible_amounts = []
    for i, line in enumerate(lines):
        line_lower = line.lower()
        for keyword in amount_keywords:
            if keyword in line_lower:
                # Search for amount on the same line
                match = re.search(amount_regex, line)
                if match:
                    amount_str = match.group(1).replace(',', '') # Basic normalization
                    if '.' in amount_str and amount_str.count('.') > 1: # e.g. 1.234.567 -> invalid for this simple replace
                        amount_str = amount_str.replace('.', '', amount_str.count('.') -1) # Keep last dot
                    amount_str = amount_str.replace(',', '.') # For formats like 1.234,56 -> 1234.56
                    try:
                        possible_amounts.append(float(amount_str))
                    except ValueError:
                        pass # Could not convert
                else:
                    # Check next few lines if no amount on current line
                    for next_line_idx in range(i + 1, min(i + 3, len(lines))):
                        match_next = re.search(amount_regex, lines[next_line_idx])
                        if match_next:
                            amount_str = match_next.group(1).replace(',', '')
                            if '.' in amount_str and amount_str.count('.') > 1:
                                amount_str = amount_str.replace('.', '', amount_str.count('.') -1)
                            amount_str = amount_str.replace(',', '.')
                            try:
                                possible_amounts.append(float(amount_str))
                                break # Found amount for this keyword
                            except ValueError:
                                pass
                break # Move to next line once a keyword is processed on this line
    
    if possible_amounts:
        # Often the grand total is the largest amount found, but this is heuristic
        invoice_data["total_amount"] = max(possible_amounts) 
        # Alternative: look for specific keywords like "grand total" and prioritize that.
        # For now, max is a simple approach.

    # --- Vendor/Client Name --- (Can be very tricky)
    # Keywords: To, Bill To, Client, Sold To, Attention, Attn, Vendor, From, Supplier
    # Often at the top, or near an address. Sometimes it's just the first prominent name.
    vendor_keywords = ["to:", "bill to:", "client:", "sold to:", "attention:", "attn:", "customer:"]
    supplier_keywords = ["from:", "vendor:", "supplier:", "invoice from:"]
    
    # Attempt to find client/customer name
    found_vendor = False
    for i, line in enumerate(lines):
        line_lower = line.lower()
        for keyword in vendor_keywords:
            if keyword in line_lower:
                # The name is often on the same line after the keyword, or the next non-empty line
                potential_name_line = line.split(keyword, 1)[-1].strip()
                if not potential_name_line and i + 1 < len(lines):
                    potential_name_line = lines[i+1].strip()
                
                if potential_name_line and len(potential_name_line) > 2: # Basic check
                    # Further refinement: avoid lines that are just addresses or dates
                    if not re.match(r"^(\d{1,4}\s|P\.?O\.?\sBox)", potential_name_line, re.IGNORECASE) and not re.search(date_patterns[0], potential_name_line, re.IGNORECASE):
                        invoice_data["vendor_name"] = potential_name_line.split('\n')[0].strip() # Take first line of it
                        found_vendor = True
                        break
        if found_vendor:
            break
    
    # If no client found with keywords, try a more general approach (e.g. first few lines)
    if not found_vendor:
        for i in range(min(5, len(lines))): # Check first 5 lines
            line_content = lines[i].strip()
            # Avoid lines that are clearly "Invoice", "Date", numbers, or too short
            if line_content and not line_content.lower().startswith(("invoice", "date", "page")) and not line_content.isdigit() and len(line_content) > 5:
                # This is a very rough heuristic, might pick up sender too.
                # A more robust solution might involve Named Entity Recognition (NER)
                # or looking for company suffixes (Ltd, Inc, LLC)
                # For now, let's assume the first suitable line could be it if not found by keywords.
                # Example: Check for company suffixes
                if re.search(r"\b(LTD|LLC|INC|LIMITED|GMBH|SARL|CORP|PLC|CO\.?)\b", line_content, re.IGNORECASE):
                     invoice_data["vendor_name"] = line_content
                     break
                # If still no vendor_name, and we haven't found one via keywords, this is a fallback
                if not invoice_data["vendor_name"] and i < 2: # Only take from first 2 lines as a last resort fallback
                    invoice_data["vendor_name"] = line_content
                    # break # Commented out to allow company suffix check to override if it comes later in first 5 lines

    print(f"Parsed data: {invoice_data}")
    return invoice_data

def write_to_excel(all_invoice_data, output_file, existing_excel_path=None):
    """
    Writes the extracted invoice data to an Excel file.
    If existing_excel_path is provided, it appends to that file.
    Otherwise, it creates a new file.
    """
    if existing_excel_path and os.path.exists(existing_excel_path):
        try:
            workbook = openpyxl.load_workbook(existing_excel_path)
            sheet = workbook.active
            # Check if headers are already present, if not, add them (optional, good practice)
            # For simplicity, we assume if the file exists and is loaded, we just append data.
            # A more robust check would verify header row.
            print(f"Appending data to existing Excel file: {existing_excel_path}")
        except Exception as e:
            print(f"Error loading existing Excel file {existing_excel_path}: {e}. Creating a new one instead at {output_file}")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = ["Invoice Number", "Invoice Date", "Total Amount", "Vendor Name", "File Name"]
            sheet.append(headers)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Invoice Number", "Invoice Date", "Total Amount", "Vendor Name", "File Name"]
        sheet.append(headers)
        if existing_excel_path:
            print(f"Existing Excel file not found at {existing_excel_path}. Creating a new one at {output_file}")
        else:
            print(f"Creating new Excel file: {output_file}")

    for invoice_data in all_invoice_data:
        row = [
            invoice_data.get("invoice_number", ""),
            invoice_data.get("invoice_date", ""),
            invoice_data.get("total_amount", ""),
            invoice_data.get("vendor_name", ""),
            invoice_data.get("file_name", ""),
        ]
        sheet.append(row)
    
    workbook.save(output_file)
    print(f"Data written to Excel file: {output_file}")

# --- Main Processing Logic ---

def process_invoices():
    """
    Main function to process all invoices in the directory.
    """
    if not os.path.exists(INVOICE_DIR):
        print(f"Error: Invoice directory '{INVOICE_DIR}' not found. Please create it and add your invoice files.")
        return

    all_extracted_data = []
    
    print(f"Looking for invoices in: {os.path.abspath(INVOICE_DIR)}")
    invoice_files = os.listdir(INVOICE_DIR)
    print(f"Files found in '{INVOICE_DIR}': {invoice_files}")

    for filename in invoice_files:
        file_path = os.path.join(INVOICE_DIR, filename)
        extracted_text = None  # Initialize for each file

        if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif')):
            print(f"\nProcessing image file: {file_path}")
            extracted_text = ocr_image(file_path)
        elif filename.lower().endswith('.pdf'):
            print(f"\nProcessing PDF file: {file_path}")
            extracted_text = extract_text_from_pdf(file_path)
            if not extracted_text or len(extracted_text.strip()) < 50: 
                print(f"Direct text extraction from {file_path} was minimal or empty. Attempting OCR by converting PDF to images.")
                extracted_text = ocr_pdf_as_images(file_path)
        elif filename.lower().endswith('.txt'):
            print(f"\nProcessing TXT file: {file_path}")
            extracted_text = read_text_from_txt(file_path)
        else:
            print(f"Skipping unsupported file type: {filename}")
            continue
            
        if extracted_text and extracted_text.strip():
            print(f"--- Raw Extracted Text for {filename} (first 300 chars) ---")
            print(extracted_text[:300] + ("..." if len(extracted_text) > 300 else ""))
            print("--- End Raw Extracted Text ---")
            invoice_details = parse_invoice_text(extracted_text)
            invoice_details["file_name"] = filename
            all_extracted_data.append(invoice_details)
        elif extracted_text is not None:
            print(f"No text could be extracted from {filename}.")
            
    if all_extracted_data:
        write_to_excel(all_extracted_data, OUTPUT_EXCEL_FILE)
    else:
        print("No data successfully extracted and parsed from any invoices.")

if __name__ == "__main__":
    if not os.path.exists(INVOICE_DIR):
        os.makedirs(INVOICE_DIR)
        print(f"Created directory: {INVOICE_DIR}. Please add your invoice files there.")
    process_invoices()
